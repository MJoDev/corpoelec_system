import sqlite3
import os
import sys
import shutil
import base64
import mimetypes
import threading
import webbrowser
from datetime import datetime
from flask import Flask, jsonify, request, send_from_directory, abort

app = Flask(__name__, static_folder=None)

# ─── Paths ────────────────────────────────────────────────────────────────────

def _app_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def _static_dir():
    if getattr(sys, 'frozen', False):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))

APP_DIR   = _app_dir()
STATIC    = _static_dir()
DB_PATH   = os.path.join(APP_DIR, 'sgip_data.db')
ASSETS    = os.path.join(APP_DIR, 'assets', 'images')
HTML_PATH = os.path.join(STATIC, 'sgip.html')
PORT      = 5128

# ─── Database helpers ─────────────────────────────────────────────────────────

def _conn():
    c = sqlite3.connect(DB_PATH)
    c.row_factory = sqlite3.Row
    c.execute("PRAGMA foreign_keys = ON")
    return c

def _init_db():
    os.makedirs(ASSETS, exist_ok=True)
    with _conn() as c:
        c.executescript("""
        CREATE TABLE IF NOT EXISTS Ubicaciones (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            planta_sede      TEXT NOT NULL,
            modulo_pasillo   TEXT NOT NULL,
            estanteria       TEXT NOT NULL,
            nivel_ubicacion  TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS Materiales (
            id                INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo            TEXT UNIQUE NOT NULL,
            nombre            TEXT NOT NULL,
            descripcion_tecnica TEXT,
            categoria         TEXT,
            estado            TEXT DEFAULT 'Bueno'
                              CHECK(estado IN ('Bueno','Dañado','En Reparación')),
            stock_actual      INTEGER DEFAULT 0,
            stock_minimo      INTEGER DEFAULT 0,
            ruta_imagen       TEXT,
            ubicacion_id      INTEGER REFERENCES Ubicaciones(id)
        );
        CREATE TABLE IF NOT EXISTS Movimientos_Kardex (
            id                   INTEGER PRIMARY KEY AUTOINCREMENT,
            material_id          INTEGER REFERENCES Materiales(id),
            tipo_movimiento      TEXT
                                 CHECK(tipo_movimiento IN ('Entrada','Salida','Transferencia')),
            cantidad             INTEGER,
            fecha_hora           DATETIME DEFAULT CURRENT_TIMESTAMP,
            responsable_retiro   TEXT,
            orden_trabajo_area   TEXT,
            ubicacion_origen_id  INTEGER REFERENCES Ubicaciones(id),
            ubicacion_destino_id INTEGER REFERENCES Ubicaciones(id)
        );
        """)

        for col in ("stock_nuevo", "stock_usado", "stock_danado"):
            try:
                c.execute("ALTER TABLE Materiales ADD COLUMN {0} INTEGER DEFAULT 0".format(col))
            except sqlite3.OperationalError:
                pass

        try:
            c.execute("ALTER TABLE Movimientos_Kardex ADD COLUMN condicion TEXT")
        except sqlite3.OperationalError:
            pass

        c.execute("""
            UPDATE Materiales SET stock_nuevo = stock_actual
            WHERE stock_nuevo=0 AND stock_usado=0 AND stock_danado=0
              AND estado IN ('Bueno','En Reparación') AND stock_actual > 0
        """)
        c.execute("""
            UPDATE Materiales SET stock_danado = stock_actual
            WHERE stock_nuevo=0 AND stock_usado=0 AND stock_danado=0
              AND estado = 'Dañado' AND stock_actual > 0
        """)
        c.commit()

def _derive_estado(stock_nuevo, stock_usado, stock_danado):
    if stock_danado > 0:
        return "Dañado"
    return "Bueno"

def _img_to_data_url(path):
    mime, _ = mimetypes.guess_type(path)
    mime = mime or 'image/jpeg'
    with open(path, 'rb') as f:
        b64 = base64.b64encode(f.read()).decode('ascii')
    return 'data:{0};base64,{1}'.format(mime, b64)

def _file_dialog(title, filetypes):
    """Open a native file dialog; returns path string or None."""
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.wm_attributes('-topmost', 1)
        path = filedialog.askopenfilename(title=title, filetypes=filetypes)
        root.destroy()
        return path or None
    except Exception:
        return None

# ─── Static files ─────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return send_from_directory(STATIC, 'sgip.html')

@app.route('/static/<path:filename>')
def static_files(filename):
    return send_from_directory(os.path.join(STATIC, 'static'), filename)

# ─── Ubicaciones ──────────────────────────────────────────────────────────────

@app.route('/api/get_ubicaciones')
def get_ubicaciones():
    with _conn() as c:
        rows = c.execute(
            "SELECT * FROM Ubicaciones ORDER BY planta_sede, modulo_pasillo, estanteria, nivel_ubicacion"
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/crear_ubicacion', methods=['POST'])
def crear_ubicacion():
    d = request.get_json()
    with _conn() as c:
        cur = c.execute(
            "INSERT INTO Ubicaciones(planta_sede, modulo_pasillo, estanteria, nivel_ubicacion) VALUES(?,?,?,?)",
            (d['planta_sede'], d['modulo_pasillo'], d['estanteria'], d['nivel_ubicacion']))
        c.commit()
    return jsonify({"ok": True, "id": cur.lastrowid})

@app.route('/api/editar_ubicacion', methods=['POST'])
def editar_ubicacion():
    d = request.get_json()
    with _conn() as c:
        c.execute(
            "UPDATE Ubicaciones SET planta_sede=?, modulo_pasillo=?, estanteria=?, nivel_ubicacion=? WHERE id=?",
            (d['planta_sede'], d['modulo_pasillo'], d['estanteria'], d['nivel_ubicacion'], d['ubicacion_id']))
        c.commit()
    return jsonify({"ok": True})

@app.route('/api/eliminar_ubicacion', methods=['POST'])
def eliminar_ubicacion():
    d = request.get_json()
    ubicacion_id = d['ubicacion_id']
    with _conn() as c:
        in_use = c.execute(
            "SELECT COUNT(*) FROM Materiales WHERE ubicacion_id=?", (ubicacion_id,)).fetchone()[0]
    if in_use:
        return jsonify({"ok": False, "error": "Ubicación en uso por materiales existentes"})
    with _conn() as c:
        c.execute("DELETE FROM Ubicaciones WHERE id=?", (ubicacion_id,))
        c.commit()
    return jsonify({"ok": True})

# ─── Materiales ───────────────────────────────────────────────────────────────

@app.route('/api/get_materiales')
def get_materiales():
    with _conn() as c:
        rows = c.execute("""
            SELECT m.*,
                   u.planta_sede, u.modulo_pasillo, u.estanteria, u.nivel_ubicacion
            FROM   Materiales m
            LEFT JOIN Ubicaciones u ON m.ubicacion_id = u.id
            ORDER  BY m.nombre
        """).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d['bajo_minimo'] = d['stock_actual'] <= d['stock_minimo']
        if d.get('ruta_imagen'):
            abs_path = os.path.join(APP_DIR, d['ruta_imagen'].replace('/', os.sep))
            if os.path.isfile(abs_path):
                d['ruta_imagen'] = _img_to_data_url(abs_path)
            else:
                d['ruta_imagen'] = None
        result.append(d)
    return jsonify(result)

@app.route('/api/crear_material', methods=['POST'])
def crear_material():
    d = request.get_json()
    sn = int(d['stock_nuevo'])
    su = int(d['stock_usado'])
    sd = int(d['stock_danado'])
    stock_actual = sn + su + sd
    estado = _derive_estado(sn, su, sd)
    with _conn() as c:
        try:
            cur = c.execute("""
                INSERT INTO Materiales(codigo, nombre, descripcion_tecnica, categoria,
                                      estado, stock_actual, stock_nuevo, stock_usado,
                                      stock_danado, stock_minimo, ubicacion_id)
                VALUES(?,?,?,?,?,?,?,?,?,?,?)
            """, (d['codigo'], d['nombre'], d.get('descripcion_tecnica'), d.get('categoria'),
                  estado, stock_actual, sn, su, sd,
                  int(d.get('stock_minimo', 0)),
                  d['ubicacion_id'] if d.get('ubicacion_id') else None))
            c.commit()
            return jsonify({"ok": True, "id": cur.lastrowid})
        except sqlite3.IntegrityError:
            return jsonify({"ok": False, "error": "El código '{0}' ya existe".format(d['codigo'])})

@app.route('/api/editar_material', methods=['POST'])
def editar_material():
    d = request.get_json()
    sn = int(d['stock_nuevo'])
    su = int(d['stock_usado'])
    sd = int(d['stock_danado'])
    stock_actual = sn + su + sd
    estado = _derive_estado(sn, su, sd)
    with _conn() as c:
        c.execute("""
            UPDATE Materiales
            SET    nombre=?, descripcion_tecnica=?, categoria=?, estado=?,
                   stock_actual=?, stock_nuevo=?, stock_usado=?, stock_danado=?,
                   stock_minimo=?, ubicacion_id=?
            WHERE  id=?
        """, (d['nombre'], d.get('descripcion_tecnica'), d.get('categoria'), estado,
              stock_actual, sn, su, sd,
              int(d.get('stock_minimo', 0)),
              d['ubicacion_id'] if d.get('ubicacion_id') else None,
              d['material_id']))
        c.commit()
    return jsonify({"ok": True})

@app.route('/api/eliminar_material', methods=['POST'])
def eliminar_material():
    d = request.get_json()
    material_id = d['material_id']
    with _conn() as c:
        c.execute("DELETE FROM Movimientos_Kardex WHERE material_id=?", (material_id,))
        c.execute("DELETE FROM Materiales WHERE id=?", (material_id,))
        c.commit()
    return jsonify({"ok": True})

@app.route('/api/preview_imagen')
def preview_imagen():
    path = request.args.get('path', '')
    if not path or not os.path.isfile(path):
        return jsonify({"ok": False})
    try:
        return jsonify({"ok": True, "data": _img_to_data_url(path)})
    except Exception:
        return jsonify({"ok": False})

@app.route('/api/seleccionar_imagen')
def seleccionar_imagen():
    path = _file_dialog(
        "Seleccionar imagen",
        [('Imágenes', '*.png *.jpg *.jpeg *.gif *.bmp *.webp'), ('Todos', '*.*')]
    )
    if path:
        return jsonify({"ok": True, "path": path})
    return jsonify({"ok": False})

@app.route('/api/subir_imagen', methods=['POST'])
def subir_imagen():
    d = request.get_json()
    material_id = d['material_id']
    src_path = d['src_path']
    if not os.path.isfile(src_path):
        return jsonify({"ok": False, "error": "Archivo no encontrado"})
    ext = os.path.splitext(src_path)[1].lower()
    fname = "mat_{0}{1}".format(material_id, ext)
    dst = os.path.join(ASSETS, fname)
    shutil.copy2(src_path, dst)
    rel = os.path.join("assets", "images", fname).replace("\\", "/")
    with _conn() as c:
        c.execute("UPDATE Materiales SET ruta_imagen=? WHERE id=?", (rel, material_id))
        c.commit()
    return jsonify({"ok": True, "ruta": _img_to_data_url(dst)})

# ─── Movimientos / Kárdex ────────────────────────────────────────────────────

@app.route('/api/get_movimientos')
def get_movimientos():
    material_id = request.args.get('material_id')
    desde       = request.args.get('desde')
    hasta       = request.args.get('hasta')
    tipo        = request.args.get('tipo')

    where, params = [], []
    if material_id:
        where.append("mk.material_id = ?"); params.append(material_id)
    if tipo:
        where.append("mk.tipo_movimiento = ?"); params.append(tipo)
    if desde:
        where.append("mk.fecha_hora >= ?"); params.append(desde)
    if hasta:
        where.append("mk.fecha_hora <= ?"); params.append(hasta + " 23:59:59")
    clause = ("WHERE " + " AND ".join(where)) if where else ""

    with _conn() as c:
        rows = c.execute("""
            SELECT mk.*,
                   m.codigo, m.nombre,
                   uo.planta_sede || ' / ' || uo.modulo_pasillo || ' / ' || uo.estanteria || ' / ' || uo.nivel_ubicacion AS origen_txt,
                   ud.planta_sede || ' / ' || ud.modulo_pasillo || ' / ' || ud.estanteria || ' / ' || ud.nivel_ubicacion AS destino_txt
            FROM   Movimientos_Kardex mk
            JOIN   Materiales m ON mk.material_id = m.id
            LEFT JOIN Ubicaciones uo ON mk.ubicacion_origen_id = uo.id
            LEFT JOIN Ubicaciones ud ON mk.ubicacion_destino_id = ud.id
            {0}
            ORDER  BY mk.fecha_hora DESC
        """.format(clause), params).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/registrar_movimiento', methods=['POST'])
def registrar_movimiento():
    d           = request.get_json()
    cantidad    = int(d['cantidad'])
    material_id = int(d['material_id'])
    tipo        = d['tipo_movimiento']
    condicion   = d.get('condicion')
    fecha       = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    col_map = {"Nuevo": "stock_nuevo", "Usado": "stock_usado", "Dañado": "stock_danado"}

    with _conn() as c:
        mat = c.execute(
            "SELECT stock_actual, stock_nuevo, stock_usado, stock_danado, stock_minimo, ubicacion_id FROM Materiales WHERE id=?",
            (material_id,)).fetchone()
        if not mat:
            return jsonify({"ok": False, "error": "Material no encontrado"})

        sn = mat["stock_nuevo"]
        su = mat["stock_usado"]
        sd = mat["stock_danado"]
        ub_actual = mat["ubicacion_id"]

        if tipo == "Entrada":
            col = col_map.get(condicion, "stock_nuevo")
            new_vals = {"stock_nuevo": sn, "stock_usado": su, "stock_danado": sd}
            new_vals[col] += cantidad
            nuevo_sn, nuevo_su, nuevo_sd = new_vals["stock_nuevo"], new_vals["stock_usado"], new_vals["stock_danado"]
            nuevo_stock = nuevo_sn + nuevo_su + nuevo_sd
            estado = _derive_estado(nuevo_sn, nuevo_su, nuevo_sd)
            c.execute("""UPDATE Materiales
                         SET stock_nuevo=?, stock_usado=?, stock_danado=?,
                             stock_actual=?, estado=?
                         WHERE id=?""",
                      (nuevo_sn, nuevo_su, nuevo_sd, nuevo_stock, estado, material_id))

        elif tipo == "Salida":
            col = col_map.get(condicion, "stock_nuevo")
            stock_col = mat[col]
            if cantidad > stock_col:
                label = condicion or "total"
                return jsonify({"ok": False, "error": "Stock {0} insuficiente (disponible: {1})".format(label, stock_col)})
            new_vals = {"stock_nuevo": sn, "stock_usado": su, "stock_danado": sd}
            new_vals[col] -= cantidad
            nuevo_sn, nuevo_su, nuevo_sd = new_vals["stock_nuevo"], new_vals["stock_usado"], new_vals["stock_danado"]
            nuevo_stock = nuevo_sn + nuevo_su + nuevo_sd
            estado = _derive_estado(nuevo_sn, nuevo_su, nuevo_sd)
            c.execute("""UPDATE Materiales
                         SET stock_nuevo=?, stock_usado=?, stock_danado=?,
                             stock_actual=?, estado=?
                         WHERE id=?""",
                      (nuevo_sn, nuevo_su, nuevo_sd, nuevo_stock, estado, material_id))

        elif tipo == "Transferencia":
            dest_id = d.get('ubicacion_destino_id')
            if not dest_id:
                return jsonify({"ok": False, "error": "Se requiere ubicación destino"})
            c.execute("UPDATE Materiales SET ubicacion_id=? WHERE id=?",
                      (int(dest_id), material_id))
            nuevo_stock = mat["stock_actual"]
            nuevo_sn, nuevo_su, nuevo_sd = sn, su, sd
        else:
            return jsonify({"ok": False, "error": "Tipo de movimiento inválido"})

        orig = int(d['ubicacion_origen_id']) if d.get('ubicacion_origen_id') else (ub_actual if tipo == "Transferencia" else None)
        dest = int(d['ubicacion_destino_id']) if d.get('ubicacion_destino_id') else None

        cur = c.execute("""
            INSERT INTO Movimientos_Kardex
                (material_id, tipo_movimiento, cantidad, fecha_hora,
                 responsable_retiro, orden_trabajo_area,
                 ubicacion_origen_id, ubicacion_destino_id, condicion)
            VALUES(?,?,?,?,?,?,?,?,?)
        """, (material_id, tipo, cantidad, fecha,
              d.get('responsable_retiro'), d.get('orden_trabajo_area'),
              orig, dest, condicion))
        c.commit()

    with _conn() as c:
        m2 = c.execute("SELECT stock_minimo FROM Materiales WHERE id=?", (material_id,)).fetchone()
    bajo = nuevo_stock <= m2["stock_minimo"] if m2 else False

    return jsonify({"ok": True, "id": cur.lastrowid, "nuevo_stock": nuevo_stock, "bajo_minimo": bajo})

# ─── Exportaciones ────────────────────────────────────────────────────────────

@app.route('/api/exportar_excel', methods=['POST'])
def exportar_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return jsonify({"ok": False, "error": "openpyxl no instalado"})

    try:
        return _hacer_excel(request.get_json())
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

def _hacer_excel(d):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    tipo = d.get('tipo', 'inventario')
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = os.path.join(APP_DIR, "SGIP_{0}_{1}.xlsx".format(tipo, ts))

    wb = openpyxl.Workbook()
    ws = wb.active
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(fill_type="solid", fgColor="2D4A8A")

    if tipo == "inventario":
        ws.title = "Existencias"
        headers = ["Código", "Nombre", "Categoría",
                   "Stock Nuevo", "Stock Usado", "Stock Dañado", "Stock Total",
                   "Stock Mínimo", "Ubicación"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font; cell.fill = hdr_fill
        with _conn() as c:
            mats = c.execute("""
                SELECT m.*, u.planta_sede, u.modulo_pasillo, u.estanteria, u.nivel_ubicacion
                FROM Materiales m LEFT JOIN Ubicaciones u ON m.ubicacion_id = u.id
                ORDER BY m.nombre
            """).fetchall()
        for m in mats:
            m = dict(m)
            ub = " / ".join(filter(None, [
                m.get("planta_sede"), m.get("modulo_pasillo"),
                m.get("estanteria"), m.get("nivel_ubicacion")]))
            ws.append([m["codigo"], m["nombre"], m["categoria"],
                       m.get("stock_nuevo", 0), m.get("stock_usado", 0), m.get("stock_danado", 0),
                       m["stock_actual"], m["stock_minimo"], ub])
    else:
        ws.title = "Kárdex"
        headers = ["ID", "Fecha/Hora", "Tipo", "Condición", "Código", "Material",
                   "Cantidad", "Responsable", "OT/Área", "Origen", "Destino"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font; cell.fill = hdr_fill
        desde = d.get('desde')
        hasta = d.get('hasta')
        where, params = [], []
        if desde:
            where.append("mk.fecha_hora >= ?"); params.append(desde)
        if hasta:
            where.append("mk.fecha_hora <= ?"); params.append(hasta + " 23:59:59")
        clause = ("WHERE " + " AND ".join(where)) if where else ""
        with _conn() as c:
            movs = c.execute("""
                SELECT mk.*, m.codigo, m.nombre,
                       uo.planta_sede||' / '||uo.modulo_pasillo AS origen_txt,
                       ud.planta_sede||' / '||ud.modulo_pasillo AS destino_txt
                FROM Movimientos_Kardex mk
                JOIN Materiales m ON mk.material_id = m.id
                LEFT JOIN Ubicaciones uo ON mk.ubicacion_origen_id = uo.id
                LEFT JOIN Ubicaciones ud ON mk.ubicacion_destino_id = ud.id
                {0} ORDER BY mk.fecha_hora DESC
            """.format(clause), params).fetchall()
        for m in movs:
            m = dict(m)
            ws.append([m["id"], m["fecha_hora"], m["tipo_movimiento"],
                       m.get("condicion", ""), m["codigo"], m["nombre"], m["cantidad"],
                       m["responsable_retiro"], m["orden_trabajo_area"],
                       m.get("origen_txt", ""), m.get("destino_txt", "")])

    for col in ws.columns:
        cell = col[0]
        letter = cell.column_letter if hasattr(cell, 'column_letter') else cell.column
        ws.column_dimensions[letter].width = 18
    wb.save(dest)
    return jsonify({"ok": True, "path": dest})

@app.route('/api/exportar_pdf', methods=['POST'])
def exportar_pdf():
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
    except ImportError:
        return jsonify({"ok": False, "error": "reportlab no instalado"})

    d    = request.get_json()
    tipo = d.get('tipo', 'inventario')
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = os.path.join(APP_DIR, "SGIP_{0}_{1}.pdf".format(tipo, ts))

    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(dest, pagesize=landscape(A4),
                            leftMargin=30, rightMargin=30,
                            topMargin=30, bottomMargin=30)
    story = []

    if tipo == "inventario":
        story.append(Paragraph("SGIP — Reporte de Existencias", styles["Title"]))
        story.append(Paragraph(datetime.now().strftime("%d/%m/%Y %H:%M"), styles["Normal"]))
        story.append(Spacer(1, 12))
        headers = [["Código", "Nombre", "Categoría", "Nuevo", "Usado", "Dañado", "Total", "Mínimo", "Ubicación"]]
        with _conn() as c:
            mats = c.execute("""
                SELECT m.*, u.modulo_pasillo, u.estanteria
                FROM Materiales m LEFT JOIN Ubicaciones u ON m.ubicacion_id = u.id
                ORDER BY m.nombre
            """).fetchall()
        data = headers + [
            [dict(m)["codigo"], dict(m)["nombre"][:28], dict(m)["categoria"],
             str(dict(m).get("stock_nuevo", 0)), str(dict(m).get("stock_usado", 0)),
             str(dict(m).get("stock_danado", 0)), str(dict(m)["stock_actual"]),
             str(dict(m)["stock_minimo"]),
             "{0}/{1}".format(dict(m).get('modulo_pasillo',''), dict(m).get('estanteria',''))]
            for m in mats
        ]
    else:
        story.append(Paragraph("SGIP — Kárdex de Movimientos", styles["Title"]))
        story.append(Paragraph(datetime.now().strftime("%d/%m/%Y %H:%M"), styles["Normal"]))
        story.append(Spacer(1, 12))
        headers = [["ID", "Fecha", "Tipo", "Condición", "Código", "Material", "Cant.", "Responsable"]]
        desde = d.get('desde')
        hasta = d.get('hasta')
        where, params = [], []
        if desde:
            where.append("mk.fecha_hora >= ?"); params.append(desde)
        if hasta:
            where.append("mk.fecha_hora <= ?"); params.append(hasta + " 23:59:59")
        clause = ("WHERE " + " AND ".join(where)) if where else ""
        with _conn() as c:
            movs = c.execute("""
                SELECT mk.*, m.codigo, m.nombre
                FROM Movimientos_Kardex mk
                JOIN Materiales m ON mk.material_id = m.id
                {0} ORDER BY mk.fecha_hora DESC
            """.format(clause), params).fetchall()
        data = headers + [
            [str(dict(m)["id"]), dict(m)["fecha_hora"][:16], dict(m)["tipo_movimiento"],
             dict(m).get("condicion", "") or "—",
             dict(m)["codigo"], dict(m)["nombre"][:22], str(dict(m)["cantidad"]),
             dict(m)["responsable_retiro"] or ""]
            for m in movs
        ]

    tbl = Table(data)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2D4A8A")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4FF")]),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("ALIGN",      (0, 0), (-1, -1), "LEFT"),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(tbl)
    doc.build(story)
    return jsonify({"ok": True, "path": dest})

# ─── Config / Info ────────────────────────────────────────────────────────────

@app.route('/api/get_db_info')
def get_db_info():
    size = os.path.getsize(DB_PATH) if os.path.exists(DB_PATH) else 0
    img_count = 0
    img_size  = 0
    if os.path.exists(ASSETS):
        for f in os.listdir(ASSETS):
            fp = os.path.join(ASSETS, f)
            if os.path.isfile(fp):
                img_count += 1
                img_size  += os.path.getsize(fp)
    with _conn() as c:
        mat_count = c.execute("SELECT COUNT(*) FROM Materiales").fetchone()[0]
        mov_count = c.execute("SELECT COUNT(*) FROM Movimientos_Kardex").fetchone()[0]
    return jsonify({
        "db_path":     DB_PATH,
        "db_size_kb":  round(size / 1024, 1),
        "assets_path": ASSETS,
        "img_count":   img_count,
        "img_size_mb": round(img_size / 1024 / 1024, 1),
        "mat_count":   mat_count,
        "mov_count":   mov_count,
        "version":     "1.0.0",
    })

@app.route('/api/respaldar_db', methods=['POST'])
def respaldar_db():
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = os.path.join(APP_DIR, "sgip_backup_{0}.db".format(ts))
    shutil.copy2(DB_PATH, dest)
    return jsonify({"ok": True, "path": dest})

@app.route('/api/seleccionar_excel')
def seleccionar_excel():
    path = _file_dialog(
        "Seleccionar archivo Excel",
        [('Archivos Excel', '*.xlsx *.xls'), ('Todos', '*.*')]
    )
    if path:
        return jsonify({"ok": True, "path": path})
    return jsonify({"ok": False})

@app.route('/api/importar_excel', methods=['POST'])
def importar_excel():
    try:
        import openpyxl
        d    = request.get_json()
        path = d['path']
        wb   = openpyxl.load_workbook(path, data_only=True)
        ws   = wb.active
        importados = 0
        omitidos   = 0
        errores    = []
        ub_cache   = {}

        for row in [tuple(c.value for c in r) for r in ws.iter_rows(min_row=3)]:
            codigo = row[0]
            if not codigo:
                continue
            codigo   = str(codigo).strip()
            nombre   = str(row[3] or "").strip() or codigo
            descripcion = str(row[5] or "").strip()

            stock_usado  = int(row[8]  or 0)
            stock_danado = int(row[9]  or 0)
            stock_nuevo  = int(row[10] or 0)

            if stock_usado == 0 and stock_danado == 0 and stock_nuevo == 0:
                stock_nuevo = int(row[7] or 0)

            stock_actual = stock_nuevo + stock_usado + stock_danado
            estado = _derive_estado(stock_nuevo, stock_usado, stock_danado)

            def _loc(val, prefix, default):
                if val is None:
                    return default
                try:
                    return "{0} {1}".format(prefix, int(val))
                except (ValueError, TypeError):
                    return str(val).strip() or default

            modulo     = _loc(row[14], "Módulo",  "Módulo 1")
            estanteria = _loc(row[11], "Est.",    "Est. 1")
            fila       = _loc(row[13], "Fila",    "Fila 1")
            planta     = "Almacen"

            ub_key = (planta, modulo, estanteria, fila)
            if ub_key not in ub_cache:
                with _conn() as c:
                    r = c.execute(
                        "SELECT id FROM Ubicaciones WHERE planta_sede=? AND modulo_pasillo=? AND estanteria=? AND nivel_ubicacion=?",
                        ub_key).fetchone()
                    if r:
                        ub_cache[ub_key] = r["id"]
                    else:
                        c.execute(
                            "INSERT INTO Ubicaciones(planta_sede, modulo_pasillo, estanteria, nivel_ubicacion) VALUES(?,?,?,?)",
                            ub_key)
                        c.commit()
                        ub_cache[ub_key] = c.execute(
                            "SELECT id FROM Ubicaciones WHERE planta_sede=? AND modulo_pasillo=? AND estanteria=? AND nivel_ubicacion=?",
                            ub_key).fetchone()["id"]

            ub_id = ub_cache[ub_key]
            try:
                with _conn() as c:
                    c.execute("""
                        INSERT INTO Materiales
                            (codigo, nombre, descripcion_tecnica, categoria, estado,
                             stock_actual, stock_nuevo, stock_usado, stock_danado,
                             stock_minimo, ubicacion_id)
                        VALUES(?,?,?,?,?,?,?,?,?,?,?)
                    """, (codigo, nombre, descripcion, "General", estado,
                          stock_actual, stock_nuevo, stock_usado, stock_danado,
                          0, ub_id))
                    c.commit()
                importados += 1
            except Exception as e:
                omitidos += 1
                if len(errores) < 5:
                    errores.append("{0}: {1}".format(codigo, e))

        return jsonify({"ok": True, "importados": importados, "omitidos": omitidos, "errores": errores})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route('/api/abrir_carpeta_exportacion', methods=['POST'])
def abrir_carpeta_exportacion():
    import subprocess
    subprocess.Popen('explorer "{0}"'.format(APP_DIR))
    return jsonify({"ok": True})

# ─── Entry point ──────────────────────────────────────────────────────────────

def _run_flask():
    app.run(host='127.0.0.1', port=PORT, debug=False, use_reloader=False)

def _open_browser_now():
    webbrowser.open('http://127.0.0.1:{0}'.format(PORT))

def _show_control_window():
    import tkinter as tk
    root = tk.Tk()
    root.title("SGIP")
    root.geometry("320x110")
    root.resizable(False, False)
    try:
        root.iconbitmap(default='')
    except Exception:
        pass

    tk.Label(root, text="SGIP esta ejecutandose en el puerto {0}".format(PORT),
             pady=10, font=("Segoe UI", 9)).pack()
    tk.Label(root, text="Cierre esta ventana para detener el servidor.",
             fg="#666666", font=("Segoe UI", 8)).pack()

    btn_frame = tk.Frame(root, pady=10)
    btn_frame.pack()
    tk.Button(btn_frame, text="Abrir navegador", command=_open_browser_now,
              width=14).pack(side=tk.LEFT, padx=4)
    tk.Button(btn_frame, text="Cerrar SGIP", command=lambda: os._exit(0),
              width=14, bg="#C0392B", fg="white").pack(side=tk.LEFT, padx=4)

    root.protocol("WM_DELETE_WINDOW", lambda: os._exit(0))
    root.mainloop()

if __name__ == '__main__':
    _init_db()

    flask_thread = threading.Thread(target=_run_flask)
    flask_thread.daemon = True
    flask_thread.start()

    import time
    time.sleep(1.0)
    _open_browser_now()

    _show_control_window()
